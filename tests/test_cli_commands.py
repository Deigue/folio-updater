"""Tests for CLI commands."""

from __future__ import annotations

import logging
import random
from datetime import datetime, timedelta
from typing import TYPE_CHECKING, Any, Callable
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from cli.commands import import_cmd
from cli.main import app as cli_app
from db import db
from mock.folio_setup import ensure_folio_exists
from utils.constants import TORONTO_TZ, Column, Table

if TYPE_CHECKING:
    from contextlib import _GeneratorContextManager
    from pathlib import Path

    from click.testing import Result
    from typer import Typer

    from app.app_context import AppContext
    from utils.config import Config

# Use basic runner configuration
runner = CliRunner()

# Constants for test assertions
EXPECTED_TRANSACTION_COUNT = 2
TYPER_INVALID_COMMAND_EXIT_CODE = 2


@pytest.fixture(autouse=True)
def suppress_logging_conflicts() -> db.Generator[None, Any, None]:
    """Suppress logging to avoid stream conflicts during CLI testing."""
    # Temporarily disable all loggers to avoid stream conflicts
    logging.disable(logging.CRITICAL)
    yield
    logging.disable(logging.NOTSET)


def run_cli_with_config(
    config: Config,
    command_app: Typer,
    args: list[str] | None = None,
) -> Result:
    """Run CLI commands with proper config mocking."""
    if args is None:
        args = []

    # Mock bootstrap.reload_config to return our test config
    with patch("app.bootstrap.reload_config") as mock_reload:
        mock_reload.return_value = config
        return runner.invoke(command_app, args)


class TestDemoCommand:
    """Test the demo command functionality."""

    def test_demo_command(
        self,
        temp_config: Callable[
            ...,
            _GeneratorContextManager[AppContext, None, None],
        ],
    ) -> None:
        """Test demo command through main CLI app."""
        with temp_config() as ctx:
            config = ctx.config
            assert not config.folio_path.exists()
            assert not config.db_path.exists()
            result = run_cli_with_config(config, cli_app, ["demo"])
            assert result.exit_code == 0
            assert "Demo portfolio created successfully!" in result.stdout
            assert config.folio_path.exists()
            assert config.db_path.exists()
            with db.get_connection() as conn:
                count = db.get_row_count(conn, Table.TXNS.value)
                assert count > 0


class TestFXCommand:
    """Test the getfx command functionality."""

    @pytest.mark.no_mock_forex
    def test_getfx_command(
        self,
        temp_config: Callable[
            ...,
            _GeneratorContextManager[AppContext, None, None],
        ],
    ) -> None:
        """Test getfx command through main CLI app."""
        with temp_config() as ctx:
            config = ctx.config
            ensure_folio_exists()

            # Remove forex data from folio to simulate fresh state.
            if config.folio_path.exists():  # pragma: no branch
                workbook = load_workbook(config.folio_path)
                if config.forex_sheet() in workbook.sheetnames:  # pragma: no branch
                    workbook.remove(workbook[config.forex_sheet()])
                    workbook.save(config.folio_path)
                workbook.close()

            # Drop the FX rates table to force fresh fetch
            with db.get_connection() as conn:
                conn.execute(f"DROP TABLE IF EXISTS {Table.FX.value}")
                conn.commit()

            result = run_cli_with_config(config, cli_app, ["getfx"])
            assert result.exit_code == 0
            assert "Successfully updated" in result.stdout
            with db.get_connection() as conn:
                count = db.get_row_count(conn, Table.FX.value)
                assert count > 0


class TestImportCommand:
    """Test the import command functionality."""

    def _create_test_excel_file(
        self,
        file_path: Path,
        sheet_name: str = "Txns",
    ) -> None:
        """Create a test Excel file with sample transaction data."""
        # Use dynamic dates based on today to ensure they're always in the future
        today = datetime.now(TORONTO_TZ).date()
        date1 = (today + timedelta(days=1)).strftime("%Y-%m-%d")
        date2 = (today + timedelta(days=2)).strftime("%Y-%m-%d")
        random.seed(file_path.name)
        # Generate randomized price and unit values for unique transactions
        price1 = round(random.uniform(50.0, 500.0), 2)  # noqa: S311
        price2 = round(random.uniform(50.0, 500.0), 2)  # noqa: S311
        units1 = round(random.uniform(1.0, 100.0), 2)  # noqa: S311
        units2 = round(random.uniform(1.0, 100.0), 2)  # noqa: S311
        test_data = {
            Column.Txn.TXN_DATE.value: [date1, date2],
            Column.Txn.ACTION.value: ["BUY", "SELL"],
            Column.Txn.AMOUNT.value: [price1 * units1, price2 * units2],
            Column.Txn.CURRENCY.value: ["USD", "USD"],
            Column.Txn.PRICE.value: [price1, price2],
            Column.Txn.UNITS.value: [units1, units2],
            Column.Txn.TICKER.value: ["AAPL", "MSFT"],
            Column.Txn.ACCOUNT.value: ["TEST-ACCOUNT", "TEST-ACCOUNT"],
        }

        df = pd.DataFrame(test_data)
        df.to_excel(file_path, index=False, sheet_name=sheet_name)

    def test_import_command_default(
        self,
        temp_config: Callable[
            ...,
            _GeneratorContextManager[AppContext, None, None],
        ],
    ) -> None:
        """Test import command default behavior (import from configured folio file)."""
        with temp_config() as ctx:
            config = ctx.config

            # Create a folio file with test data
            self._create_test_excel_file(
                config.folio_path,
                config.transactions_sheet(),
            )

            # Run import command without arguments
            result = run_cli_with_config(config, import_cmd.app)
            assert result.exit_code == 0
            assert (
                f"Successfully imported {EXPECTED_TRANSACTION_COUNT} transactions"
                in result.stdout
            )

            # Verify data was imported to database
            with db.get_connection() as conn:
                count = db.get_row_count(conn, Table.TXNS.value)
                assert count == EXPECTED_TRANSACTION_COUNT

    def test_import_command_missing_folio(
        self,
        temp_config: Callable[
            ...,
            _GeneratorContextManager[AppContext, None, None],
        ],
    ) -> None:
        """Test import command when folio file doesn't exist."""
        with temp_config() as ctx:
            config = ctx.config
            assert not config.folio_path.exists()
            result = run_cli_with_config(config, import_cmd.app)
            assert result.exit_code == 1
            assert f"Folio file not found: {config.folio_path}" in result.stderr

    def test_import_command_file(
        self,
        temp_config: Callable[
            ...,
            _GeneratorContextManager[AppContext, None, None],
        ],
    ) -> None:
        """Test import command with specific file option."""
        with temp_config() as ctx:
            config = ctx.config
            test_file = config.project_root / "test_import.xlsx"
            self._create_test_excel_file(test_file)

            result = run_cli_with_config(
                config,
                import_cmd.app,
                ["--file", str(test_file)],
            )
            assert result.exit_code == 0
            # Verify via stdout keywords, core functionality is tested elsewhere.
            assert f"Importing {test_file.name}..." in result.stdout
            assert (
                f"Successfully imported {EXPECTED_TRANSACTION_COUNT} transactions "
                f"from {test_file.name}" in result.stdout
            )
            assert (
                f"Created folio Excel with {EXPECTED_TRANSACTION_COUNT} transactions"
                in result.stdout
            )
            processed_folder = config.project_root / "processed"
            assert processed_folder.exists()
            assert (processed_folder / test_file.name).exists()
            assert not test_file.exists()

    def test_import_command_directory(
        self,
        temp_config: Callable[
            ...,
            _GeneratorContextManager[AppContext, None, None],
        ],
    ) -> None:
        """Test import command with directory option."""
        with temp_config() as ctx:
            config = ctx.config

            # Create test directory with multiple files to import
            import_dir = config.project_root / "import_files"
            import_dir.mkdir()
            file1 = import_dir / "transactions1.xlsx"
            file2 = import_dir / "transactions2.xlsx"
            self._create_test_excel_file(file1)
            self._create_test_excel_file(file2)

            ensure_folio_exists()
            result = run_cli_with_config(
                config,
                import_cmd.app,
                ["--dir", str(import_dir)],
            )
            assert result.exit_code == 0
            assert "Found 2 files to import" in result.stdout
            assert "Total transactions imported: 4" in result.stdout
            assert "Export completed" in result.stdout
            processed_folder = config.project_root / "processed"
            assert processed_folder.exists()
            assert (processed_folder / file1.name).exists()
            assert (processed_folder / file2.name).exists()
            assert not file1.exists()
            assert not file2.exists()


class TestVersionCommand:
    """Test the version command functionality."""

    def test_version_command(self) -> None:
        """Test version command output."""
        result = runner.invoke(cli_app, ["version"])

        assert result.exit_code == 0
        assert "folio-updater version:" in result.stdout


class TestCliErrorHandling:
    """Test CLI error handling and edge cases."""

    def test_invalid_command(self) -> None:
        """Test handling of invalid commands."""
        result = runner.invoke(cli_app, ["invalid-command"])

        assert result.exit_code == TYPER_INVALID_COMMAND_EXIT_CODE
